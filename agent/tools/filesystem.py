"""文件系统工具：读取、写入、编辑文件，以及列举目录。"""

import base64
import asyncio
import difflib
import io
import logging
import mimetypes
import os
from pathlib import Path
from typing import Any

from agent.tools.base import Tool, ToolResult

logger = logging.getLogger(__name__)
_FILE_MUTATION_LOCKS: dict[str, asyncio.Lock] = {}


def _read_xlsx(file_path: Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        parts.append(f"## Sheet: {sheet}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                parts.append("\t".join(cells))
    text = "\n".join(parts)
    if len(text) > _XLSX_MAX_CHARS:
        text = (
            text[:_XLSX_MAX_CHARS]
            + f"\n\n[已截断：表格内容过长，仅返回前 {_XLSX_MAX_CHARS} 字符]"
        )
    return text


def _read_xls(file_path: Path) -> str:
    import xlrd

    wb = xlrd.open_workbook(str(file_path))
    parts = []
    for sheet in wb.sheets():
        parts.append(f"## Sheet: {sheet.name}")
        for row_idx in range(sheet.nrows):
            cells = [str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
            parts.append("\t".join(cells))
    text = "\n".join(parts)
    if len(text) > _XLSX_MAX_CHARS:
        text = (
            text[:_XLSX_MAX_CHARS]
            + f"\n\n[已截断：表格内容过长，仅返回前 {_XLSX_MAX_CHARS} 字符]"
        )
    return text


def _resolve_path(path: str, allowed_dir: Path | None = None) -> Path:
    """解析路径（展开 ~ 并取绝对路径），可选限制在允许目录内。

    相对路径规则：
    - 若提供了 allowed_dir，相对路径基于 allowed_dir 解析（工作目录为 allowed_dir）
    - 否则相对路径基于进程 cwd 解析
    """
    p = Path(path).expanduser()
    if not p.is_absolute() and allowed_dir is not None:
        resolved = (allowed_dir / p).resolve()
    else:
        resolved = p.resolve()
    if allowed_dir and not str(resolved).startswith(str(allowed_dir.resolve())):
        raise PermissionError(f"路径 {path} 超出允许目录 {allowed_dir}")
    return resolved


def _strip_utf8_bom(text: str) -> tuple[str, bool]:
    if text.startswith("\ufeff"):
        return text[1:], True
    return text, False


def _normalize_to_lf(text: str) -> str:
    return text.replace("\r\n", "\n").replace("\r", "\n")


def _restore_utf8_bom(text: str, has_bom: bool) -> str:
    if has_bom:
        return "\ufeff" + text
    return text


def _supports_crlf_compat(text: str) -> bool:
    if "\r\n" not in text:
        return False
    bare_lf = "\n" in text.replace("\r\n", "")
    return not bare_lf and "\r" not in text.replace("\r\n", "")


def _build_edit_diff(old_text: str, new_text: str, path: str) -> str:
    lines = list(
        difflib.unified_diff(
            old_text.splitlines(),
            new_text.splitlines(),
            fromfile=f"{path} (before)",
            tofile=f"{path} (after)",
            lineterm="",
            n=2,
        )
    )
    return "\n".join(lines)


def _get_file_mutation_key(file_path: Path) -> str:
    try:
        return str(file_path.resolve(strict=True))
    except FileNotFoundError:
        return os.path.realpath(str(file_path))


async def _run_with_file_mutation_lock(file_path: Path, fn: Any) -> Any:
    key = _get_file_mutation_key(file_path)
    lock = _FILE_MUTATION_LOCKS.get(key)
    if lock is None:
        lock = asyncio.Lock()
        _FILE_MUTATION_LOCKS[key] = lock

    async with lock:
        result = await fn()

    current = _FILE_MUTATION_LOCKS.get(key)
    if current is lock and not lock.locked():
        _FILE_MUTATION_LOCKS.pop(key, None)
    return result


_READ_MAX_CHARS = 10_000  # 默认单次最多返回 10K 字符，大文件须分页读取
_XLSX_MAX_CHARS = 30_000  # xlsx/xls 表格内容上限（表格行通常比代码行长）
_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}
_IMAGE_MAX_EDGE = 1568
_IMAGE_TARGET_B64_LEN = 8_000_000
_IMAGE_MIN_QUALITY = 45


def _encode_image_for_model(file_path: Path) -> tuple[str, str, bool]:
    raw = file_path.read_bytes()
    raw_b64 = base64.b64encode(raw).decode()
    mime, _ = mimetypes.guess_type(file_path.name)
    if mime and mime.startswith("image/") and len(raw_b64) <= _IMAGE_TARGET_B64_LEN:
        return mime, raw_b64, False

    try:
        from PIL import Image, ImageOps
    except ModuleNotFoundError as e:
        raise RuntimeError(
            "当前环境未安装 Pillow，无法压缩大图片；请安装 Pillow 后重试"
        ) from e

    with Image.open(file_path) as img:
        img = ImageOps.exif_transpose(img)
        if img.mode not in ("RGB", "L"):
            canvas = Image.new("RGB", img.size, (255, 255, 255))
            alpha = img.getchannel("A") if "A" in img.getbands() else None
            canvas.paste(img.convert("RGB"), mask=alpha)
            img = canvas
        elif img.mode == "L":
            img = img.convert("RGB")

        if max(img.size) > _IMAGE_MAX_EDGE:
            img.thumbnail((_IMAGE_MAX_EDGE, _IMAGE_MAX_EDGE))

        chosen: bytes | None = None
        for quality in (85, 75, 65, 55, _IMAGE_MIN_QUALITY):
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=quality, optimize=True)
            candidate = buf.getvalue()
            candidate_b64 = base64.b64encode(candidate).decode()
            chosen = candidate
            if len(candidate_b64) <= _IMAGE_TARGET_B64_LEN:
                return "image/jpeg", candidate_b64, True

    if chosen is None:
        raise RuntimeError("图片压缩失败")
    return "image/jpeg", base64.b64encode(chosen).decode(), True


def _read_image(file_path: Path) -> ToolResult:
    mime, b64, compressed = _encode_image_for_model(file_path)
    note = "，已自动压缩" if compressed else ""
    return ToolResult(
        text=f"[已读取图片文件 {file_path.name}{note}，图片内容已提供给多模态模型]",
        content_blocks=[
            {
                "type": "image_url",
                "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "high"},
            }
        ],
    )


class ReadFileTool(Tool):
    """读取文件内容，支持按行分页，超大文件自动截断。"""

    def __init__(self, allowed_dir: Path | None = None):
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "read_file"

    @property
    def description(self) -> str:
        return (
            "读取文件内容。文本文件输出带行号格式（如 '     1→内容'），便于 edit_file 精确定位；"
            "Excel 文件（.xlsx/.xls）返回表格文本，不带行号；"
            "图片文件会直接提供给多模态模型查看。\n"
            "默认每次最多 10K 字符；大文件须用 limit 分页，不要依赖字符截断后的自动续读。\n\n"
            "推荐策略：先 limit=50 预览文件结构，再按需读取目标行段（offset=N limit=M）。\n"
            "并行读取：可在同一次响应中同时读取多个文件，无需逐一等待。\n"
            "参数说明：offset=跳过的行数（0-based），limit=读取行数；二者仅对文本文件生效。"
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "要读取的文件路径",
                },
                "offset": {
                    "type": "integer",
                    "description": "起始行号（0-based），默认 0",
                    "minimum": 0,
                    "default": 0,
                },
                "limit": {
                    "type": "integer",
                    "description": "最多读取行数，默认不限（受 80K 字符上限约束）",
                    "minimum": 1,
                },
            },
            "required": ["path"],
        }

    async def execute(self, path: str, **kwargs: Any) -> str | ToolResult:
        offset: int = int(kwargs.get("offset", 0))
        limit: int | None = kwargs.get("limit")
        if limit is not None:
            limit = int(limit)
        try:
            file_path = _resolve_path(path, self._allowed_dir)
            if not file_path.exists():
                return f"错误：文件不存在：{path}"
            if not file_path.is_file():
                return f"错误：路径不是文件：{path}"

            suffix = file_path.suffix.lower()
            if suffix in _IMAGE_SUFFIXES:
                return _read_image(file_path)
            if suffix == ".xlsx":
                return _read_xlsx(file_path)
            if suffix == ".xls":
                return _read_xls(file_path)

            lines = file_path.read_text(encoding="utf-8").splitlines(keepends=True)
            total_lines = len(lines)

            # 按行分页
            sliced = lines[offset:] if limit is None else lines[offset : offset + limit]

            # 带行号输出（1-based 显示值，从 offset+1 开始）
            numbered_lines = []
            for i, line in enumerate(sliced, start=offset + 1):
                numbered_lines.append(f"{i:6}\u2192{line}")
            text = "".join(numbered_lines)

            # 字符上限截断（回退到完整行边界，避免截断行中间）
            truncated_by_chars = False
            if len(text) > _READ_MAX_CHARS:
                truncated_text = text[:_READ_MAX_CHARS]
                last_newline = truncated_text.rfind("\n")
                if last_newline > 0:
                    truncated_text = truncated_text[: last_newline + 1]
                text = truncated_text
                truncated_by_chars = True

            suffix_note = ""
            end_line = offset + len(sliced)
            if truncated_by_chars:
                suffix_note = (
                    f"\n\n[已截断：文件共 {total_lines} 行，内容过长。"
                    f"建议用 limit=N 分段读取，例如 offset=0 limit=100。]"
                )
            elif offset > 0 or limit is not None:
                suffix_note = (
                    f"\n\n[第 {offset + 1}–{end_line} 行 / 共 {total_lines} 行]"
                )
            elif total_lines > len(sliced):
                suffix_note = f"\n\n[共 {total_lines} 行]"

            return text + suffix_note
        except PermissionError as e:
            return f"错误：{e}"
        except Exception as e:
            return f"读取文件失败：{e}"


class WriteFileTool(Tool):
    """将内容写入文件，自动创建所需的父目录。"""

    def __init__(self, allowed_dir: Path | None = None):
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "write_file"

    @property
    def description(self) -> str:
        return (
            "将内容写入文件（完整覆盖写）。不存在的父目录自动创建。\n\n"
            "使用规则：\n"
            "- 优先使用 edit_file 修改已有文件；仅在创建新文件或完整重写时使用 write_file\n"
            "- 写入已存在的文件前，必须先用 read_file 读取当前内容，禁止盲写\n"
            "- 不得主动创建文档文件（*.md、README）除非用户明确要求\n"
            "- 写入路径须为绝对路径或相对工作目录的合法路径"
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "要写入的文件路径"},
                "content": {"type": "string", "description": "要写入的文本内容"},
            },
            "required": ["path", "content"],
        }

    async def execute(self, path: str, content: str, **kwargs: Any) -> str:
        try:
            file_path = _resolve_path(path, self._allowed_dir)

            async def _write() -> str:
                file_path.parent.mkdir(parents=True, exist_ok=True)
                file_path.write_text(content, encoding="utf-8")
                return f"已写入 {len(content)} 字节到 {path}"

            return await _run_with_file_mutation_lock(file_path, _write)
        except PermissionError as e:
            return f"错误：{e}"
        except Exception as e:
            return f"写入文件失败：{e}"

class EditFileTool(Tool):
    """精确替换文件中的指定文本片段。"""

    def __init__(self, allowed_dir: Path | None = None):
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "edit_file"

    @property
    def description(self) -> str:
        return (
            "将文件中的 old_text 精确替换为 new_text。\n\n"
            "重要：old_text 和 new_text 是文件的原始内容，不包含 read_file 输出的行号前缀。\n"
            "从 read_file 输出复制 old_text 时，必须去掉行首的 '     N→' 前缀，只保留实际文本内容。\n"
            "old_text 必须与文件内容完全一致（含缩进和换行）。"
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "要编辑的文件路径"},
                "old_text": {
                    "type": "string",
                    "description": "要查找并替换的原始文本（必须与文件内容完全一致，不含行号前缀）",
                },
                "new_text": {"type": "string", "description": "替换后的新文本"},
                "replace_all": {
                    "type": "boolean",
                    "default": False,
                    "description": (
                        "是否替换文件中所有匹配项，默认 False（只替换第一处）。"
                        "重命名变量、批量修改相同字符串时设为 true。"
                        "不确定匹配数量时先省略，收到'出现N次'警告后再决定。"
                    ),
                },
            },
            "required": ["path", "old_text", "new_text"],
        }

    async def execute(
        self, path: str, old_text: str, new_text: str, **kwargs: Any
    ) -> str:
        replace_all: bool = bool(kwargs.get("replace_all", False))
        try:
            file_path = _resolve_path(path, self._allowed_dir)

            async def _edit() -> str:
                if not file_path.exists():
                    return f"错误：文件不存在：{path}"

                raw_content = file_path.read_bytes().decode("utf-8")
                content, has_bom = _strip_utf8_bom(raw_content)
                matched_old_text = old_text
                replacement_text = new_text

                if matched_old_text not in content and _supports_crlf_compat(content):
                    compat_old_text = old_text.replace("\n", "\r\n")
                    if compat_old_text in content:
                        matched_old_text = compat_old_text
                        replacement_text = new_text.replace("\n", "\r\n")

                if matched_old_text not in content:
                    return "错误：未找到 old_text，请确保与文件内容完全一致。"

                count = content.count(matched_old_text)
                if count > 1 and not replace_all:
                    return f"警告：old_text 在文件中出现了 {count} 次。如需全部替换，设 replace_all=true；如需精确定位，请在 old_text 中包含更多上下文。"

                new_content = (
                    content.replace(matched_old_text, replacement_text)
                    if replace_all
                    else content.replace(matched_old_text, replacement_text, 1)
                )
                replaced_count = count if replace_all else 1
                diff_text = _build_edit_diff(content, new_content, path)
                restored_content = _restore_utf8_bom(new_content, has_bom)
                file_path.write_text(restored_content, encoding="utf-8")
                if diff_text:
                    return (
                        f"已成功编辑 {path}（替换 {replaced_count} 处）\n\n"
                        f"```diff\n{diff_text}\n```"
                    )
                return f"已成功编辑 {path}（替换 {replaced_count} 处）"

            return await _run_with_file_mutation_lock(file_path, _edit)
        except PermissionError as e:
            return f"错误：{e}"
        except Exception as e:
            return f"编辑文件失败：{e}"

class ListDirTool(Tool):
    """列举目录内容。"""

    def __init__(self, allowed_dir: Path | None = None):
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "list_dir"

    @property
    def description(self) -> str:
        return "列举指定目录下的文件和子目录。"

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "要列举的目录路径"}
            },
            "required": ["path"],
        }

    async def execute(self, path: str, **kwargs: Any) -> str:
        try:
            dir_path = _resolve_path(path, self._allowed_dir)
            if not dir_path.exists():
                return f"错误：目录不存在：{path}"
            if not dir_path.is_dir():
                return f"错误：路径不是目录：{path}"

            items = []
            for item in sorted(dir_path.iterdir()):
                prefix = "📁 " if item.is_dir() else "📄 "
                items.append(f"{prefix}{item.name}")

            if not items:
                return f"目录 {path} 为空"

            return "\n".join(items)
        except PermissionError as e:
            return f"错误：{e}"
        except Exception as e:
            return f"列举目录失败：{e}"
